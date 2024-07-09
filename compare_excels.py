import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys
import os
import numpy as np
from datetime import datetime

def compare_values(val1, val2):
    if pd.isna(val1) and pd.isna(val2):
        return True
    elif isinstance(val1, (int, float)) and isinstance(val2, (int, float)):
        return np.isclose(val1, val2, equal_nan=True)
    else:
        return val1 == val2

def compare_excel_files(file1, file2, output_file):
    # Load workbooks
    wb1 = load_workbook(file1, data_only=True)
    wb2 = load_workbook(file2, data_only=True)
   
    # Get sheet names
    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)
   
    # Find common sheets
    common_sheets = sheets1.intersection(sheets2)
   
    all_differences = []

    # Open output file
    with open(output_file, 'w') as f:
        # Compare each common sheet
        for sheet_name in common_sheets:
            f.write(f"\nComparing sheet: {sheet_name}\n")
           
            # Read sheets into pandas DataFrames
            df1 = pd.read_excel(file1, sheet_name=sheet_name)
            df2 = pd.read_excel(file2, sheet_name=sheet_name)
           
            # Compare shape
            if df1.shape != df2.shape:
                f.write(f"Sheets have different shapes: {df1.shape} vs {df2.shape}\n")
                continue
           
            # Compare column names
            if list(df1.columns) != list(df2.columns):
                f.write("Sheets have different column names:\n")
                f.write(f"File 1: {list(df1.columns)}\n")
                f.write(f"File 2: {list(df2.columns)}\n")
                continue
           
            # Compare data and collect differences
            differences = []
            for idx in range(len(df1)):
                for col in df1.columns:
                    val1 = df1.loc[idx, col]
                    val2 = df2.loc[idx, col]
                    if not compare_values(val1, val2):
                        differences.append({
                            'sheet': sheet_name,
                            'row': idx + 2,  # +2 because Excel is 1-indexed and has a header row
                            'column': col,
                            'file1_value': val1,
                            'file2_value': val2
                        })
           
            # Write differences
            if differences:
                f.write(f"Found {len(differences)} differences:\n")
                for diff in differences:
                    f.write(f"  Row {diff['row']}, Column '{diff['column']}':\n")
                    f.write(f"    File 1: '{diff['file1_value']}'\n")
                    f.write(f"    File 2: '{diff['file2_value']}'\n")
            else:
                f.write("Sheets are identical\n")
            
            all_differences.extend(differences)
       
        # Report on sheets that are not in both files
        only_in_file1 = sheets1 - sheets2
        only_in_file2 = sheets2 - sheets1
        if only_in_file1:
            f.write(f"\nSheets only in {file1}: {', '.join(only_in_file1)}\n")
        if only_in_file2:
            f.write(f"\nSheets only in {file2}: {', '.join(only_in_file2)}\n")
        
        f.write(f"\nTotal differences found: {len(all_differences)}\n")
    
    return all_differences

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python script.py <file1> <file2>")
        sys.exit(1)
    file1 = sys.argv[1]
    file2 = sys.argv[2]
    # Use relative paths
    current_dir = os.path.dirname(os.path.abspath(__file__))
    file1_path = os.path.join(current_dir, file1)
    file2_path = os.path.join(current_dir, file2)
    
    # Generate timestamped output file name
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(f"{current_dir}/comparisons", f"comparison_results_{timestamp}.txt")
    
    differences = compare_excel_files(file1_path, file2_path, output_file)
    print(f"Comparison complete. Results written to {output_file}")